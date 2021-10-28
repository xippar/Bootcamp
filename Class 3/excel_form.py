from openpyxl import load_workbook
import pandas as pd
import datetime

file_path = r'C:\Users\acer\Desktop\Bootcamp\Class 3\Result.xlsx'
form_data = r'C:\Users\acer\Desktop\Bootcamp\Class 3\excel_form.xlsx'

def today():
    today = datetime.datetime.now()
    return today.strftime('%d/%m/%Y')

def get_data(report_data,index=0,column='Total Qtr'):
    report = pd.read_excel(report_data,index_col=[index])
    return report[column]

def get_form(form_data,report_data,column='Report',file_name='excel_form2.xlsx'):
    form = load_workbook(form_data)
    form_report = form[column]
    put_report(form_report,report_data)
    form.save(file_name)

def put_report(form_report,report_data):
    form_report['F2'].value = today()
    i = 0
    for row in form_report.iter_rows(min_row=5,max_row=19,min_col=2,max_col=3):
        row[0].value = report_data.index[i]
        row[1].value = report_data.values[i]
        i += 1

def main():
    report_data = get_data(file_path)
    get_form(form_data,report_data)

if __name__ == '__main__':
    file_path = input('Please input your excel report path: ')
    form_data = input('Please input your form path: ')
    main()
